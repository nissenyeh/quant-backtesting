#!/usr/bin/env python3
"""
經濟週期報酬率回測腳本

根據 NBER 經濟週期數據和 Shiller S&P 指數，計算各經濟週期的報酬率統計。

輸入:
    - 美國景氣循環完整年表_NBER.xlsx: NBER 經濟週期資料
    - SP指數(Shiller數據）.csv: Shiller 通膨調整後 S&P 總報酬指數

輸出:
    - 經濟週期報酬率分析.xlsx: 包含各週期報酬率的 Excel 表格
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# 取得腳本所在目錄
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 讀取報酬率數據
sp_data = pd.read_csv(os.path.join(SCRIPT_DIR, 'input_SP指數(Shiller數據）.csv'))
sp_data['Year'] = sp_data['Date 日期(年.月)'].apply(lambda x: int(str(x).split('.')[0]))
sp_data['Month'] = sp_data['Date 日期(年.月)'].apply(lambda x: int(str(x).split('.')[1]) if '.' in str(x) else 1)
sp_data['Total_Return_Index'] = sp_data['Real Total Return Price 通膨調整+股息再投入指數'].astype(str).str.replace(',', '').astype(float)

yearly_data = sp_data[sp_data['Month'] == 12][['Year', 'Total_Return_Index']].copy()
yearly_data = yearly_data.sort_values('Year').reset_index(drop=True)
yearly_data['Return'] = yearly_data['Total_Return_Index'].pct_change() * 100
returns_dict = dict(zip(yearly_data['Year'], yearly_data['Return']))

def get_return(year):
    r = returns_dict.get(year)
    if r is not None and not pd.isna(r):
        return r
    return None

def calc_avg(years):
    vals = [get_return(y) for y in years]
    vals = [v for v in vals if v is not None]
    return sum(vals) / len(vals) if vals else None

def fmt_pct(val):
    return f"{val:.2f}%" if val is not None else "N/A"

# 定義週期數據
cycles = [
    # 週期1: 衰退期 1926-1927
    {"type": "衰退", "name": "衰退期 (1926-10 → 1927-11)：佛州房地產泡沫破裂+颶風+Fed升息", "years": [1926, 1927], "desc": ""},

    # 週期2: 擴張期 咆哮二十年代
    {"type": "擴張", "name": "咆哮二十年代 (1927-11 → 1929-08)", "years": [1928], "desc": ""},

    # 週期3: 衰退期 大蕭條
    {"type": "衰退", "name": "衰退期 (1929-08 → 1933-03)：大蕭條", "years": [1929, 1930, 1931, 1932, 1933], "desc": "股市崩盤→銀行倒閉→全面經濟崩潰"},

    # 週期4: 擴張期 新政復甦
    {"type": "擴張", "name": "新政復甦 (1933-03 → 1937-05)", "years": [1934, 1935, 1936], "desc": ""},

    # 週期5: 衰退期 1937-1938
    {"type": "衰退", "name": "衰退期 (1937-05 → 1938-06)：Fed過早緊縮+財政削減", "years": [1937, 1938], "desc": "經濟二次探底"},

    # 週期6: 擴張期 二戰擴張
    {"type": "擴張", "name": "二戰擴張 (1938-06 → 1945-02)", "years": [1939, 1940, 1941, 1942, 1943, 1944], "desc": ""},

    # 週期7: 衰退期 1945
    {"type": "衰退", "name": "衰退期 (1945-02 → 1945-10)：二戰結束", "years": [1945], "desc": "軍工轉民用過渡期"},

    # 週期8: 擴張期 戰後調整
    {"type": "擴張", "name": "戰後調整 (1945-10 → 1948-11)", "years": [1946, 1947], "desc": ""},

    # 週期9: 衰退期 1948-1949
    {"type": "衰退", "name": "衰退期 (1948-11 → 1949-10)：戰後軍費削減", "years": [1948, 1949], "desc": "Fed緊縮"},

    # 週期10: 擴張期 戰後繁榮
    {"type": "擴張", "name": "戰後繁榮 (1949-10 → 1953-07)", "years": [1950, 1951, 1952], "desc": ""},

    # 週期11: 衰退期 1953-1954
    {"type": "衰退", "name": "衰退期 (1953-07 → 1954-05)：韓戰結束", "years": [1953, 1954], "desc": "軍費削減+Fed升息"},

    # 週期12: 擴張期 艾森豪繁榮
    {"type": "擴張", "name": "艾森豪繁榮 (1954-05 → 1957-08)", "years": [1955, 1956], "desc": ""},

    # 週期13: 衰退期 1957-1958
    {"type": "衰退", "name": "衰退期 (1957-08 → 1958-04)：Fed升息", "years": [1957, 1958], "desc": "汽車鋼鐵產能過剩"},

    # 週期14: 擴張期 短期復甦
    {"type": "擴張", "name": "短期復甦 (1958-04 → 1960-04)", "years": [1959], "desc": ""},

    # 週期15: 衰退期 1960-1961
    {"type": "衰退", "name": "衰退期 (1960-04 → 1961-02)：Fed緊縮", "years": [1960, 1961], "desc": "製造業與出口走弱"},

    # 週期16: 擴張期 甘迺迪-詹森擴張
    {"type": "擴張", "name": "甘迺迪-詹森擴張 (1961-02 → 1969-12)", "years": [1962, 1963, 1964, 1965, 1966, 1967, 1968], "desc": ""},

    # 週期17: 衰退期 1969-1970
    {"type": "衰退", "name": "衰退期 (1969-12 → 1970-11)：越戰過熱", "years": [1969, 1970], "desc": "Fed升息壓制"},

    # 週期18: 擴張期 尼克森擴張
    {"type": "擴張", "name": "尼克森擴張 (1970-11 → 1973-11)", "years": [1971, 1972], "desc": ""},

    # 週期19: 衰退期 1973-1975 石油危機
    {"type": "衰退", "name": "衰退期 (1973-11 → 1975-03)：第一次石油危機", "years": [1973, 1974, 1975], "desc": "OPEC禁運+通膨失控"},

    # 週期20: 擴張期 滯脹復甦
    {"type": "擴張", "name": "滯脹復甦 (1975-03 → 1980-01)", "years": [1976, 1977, 1978, 1979], "desc": ""},

    # 週期21: 衰退期 1980
    {"type": "衰退", "name": "衰退期 (1980-01 → 1980-07)：Volcker升息", "years": [1980], "desc": "暴力升息(17%+)打通膨，短暫衰退"},

    # 週期22: 衰退期 1981-1982 (短暫復甦只有幾個月，直接跳到下個衰退)
    {"type": "衰退", "name": "衰退期 (1981-07 → 1982-11)：Volcker再升息", "years": [1981, 1982], "desc": "升息20%，徹底壓制通膨"},

    # 週期23: 擴張期 雷根牛市
    {"type": "擴張", "name": "雷根牛市 (1982-11 → 1990-07)", "years": [1983, 1984, 1985, 1986, 1987, 1988, 1989], "desc": ""},

    # 週期24: 衰退期 1990-1991
    {"type": "衰退", "name": "衰退期 (1990-07 → 1991-03)：儲貸危機", "years": [1990, 1991], "desc": "房產泡沫+波灣戰爭"},

    # 週期25: 擴張期 克林頓繁榮
    {"type": "擴張", "name": "克林頓繁榮 (1991-03 → 2001-03)", "years": [1992, 1993, 1994, 1995, 1996, 1997, 1998, 1999, 2000], "desc": ""},

    # 週期26: 衰退期 2001
    {"type": "衰退", "name": "衰退期 (2001-03 → 2001-11)：網路泡沫破裂", "years": [2001], "desc": "911恐攻"},

    # 週期27: 擴張期 房地產泡沫
    {"type": "擴張", "name": "房地產泡沫 (2001-11 → 2007-12)", "years": [2002, 2003, 2004, 2005, 2006], "desc": ""},

    # 週期28: 衰退期 2007-2009 金融海嘯
    {"type": "衰退", "name": "衰退期 (2007-12 → 2009-06)：次貸危機", "years": [2007, 2008, 2009], "desc": "金融海嘯：雷曼倒閉，S&P500跌57%"},

    # 週期29: 擴張期 QE牛市
    {"type": "擴張", "name": "QE牛市 (2009-06 → 2020-02)", "years": [2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019], "desc": ""},

    # 週期30: 衰退期 2020 COVID
    {"type": "衰退", "name": "衰退期 (2020-02 → 2020-04)：COVID-19", "years": [2020], "desc": "史上最短衰退"},

    # 週期31: 擴張期 疫後復甦
    {"type": "擴張", "name": "疫後復甦 (2020-04 → 持續中)", "years": [2021, 2022], "desc": "持續中"},
]

# 生成表格數據
data = []
for cycle_idx, cycle in enumerate(cycles):
    cycle_num = cycle_idx + 1
    years = cycle["years"]

    for i, year in enumerate(years):
        row = {
            "週期": cycle["name"] if i == 0 else "",
            "年份": year,
            "報酬率": fmt_pct(get_return(year)),
            "定義": "擴張期" if cycle["type"] == "擴張" else "衰退期",
            "附註": "",
            "cycle_num": cycle_num
        }

        # 擴張期第一年：計算統計數據
        if cycle["type"] == "擴張" and i == 0:
            exp_avg = calc_avg(years)

            # 衰退前1年 = 擴張期最後1年
            last1 = calc_avg([years[-1]]) if len(years) >= 1 else None
            # 衰退前2年 = 擴張期最後2年平均
            last2 = calc_avg(years[-2:]) if len(years) >= 2 else None
            # 衰退前3年 = 擴張期最後3年平均
            last3 = calc_avg(years[-3:]) if len(years) >= 3 else None

            notes = []
            notes.append(f"擴張期（共{len(years)}年）平均報酬率：{fmt_pct(exp_avg)}")

            # 扣除倒數1年
            if len(years) >= 2:
                exclude_last1 = calc_avg(years[:-1])
                notes.append(f"擴張期（扣除倒數1年，共{len(years)-1}年）平均報酬率：{fmt_pct(exclude_last1)}")

            # 扣除倒數2年
            if len(years) >= 3:
                exclude_last2 = calc_avg(years[:-2])
                notes.append(f"擴張期（扣除倒數2年，共{len(years)-2}年）平均報酬率：{fmt_pct(exclude_last2)}")

            if last1 is not None:
                notes.append(f"衰退前1年報酬率：{fmt_pct(last1)}")
            if last2 is not None:
                notes.append(f"衰退前2年平均報酬率：{fmt_pct(last2)}")
            if last3 is not None:
                notes.append(f"衰退前3年平均報酬率：{fmt_pct(last3)}")

            row["附註"] = "\n".join(notes)

        # 衰退期第一年：加入描述
        elif cycle["type"] == "衰退" and i == 0 and cycle["desc"]:
            row["附註"] = cycle["desc"]

        data.append(row)

# 創建 workbook
wb = Workbook()
ws = wb.active
ws.title = "經濟週期報酬率"

# 設定樣式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 寫入標題
headers = ["週期", "年份", "報酬率", "定義", "附註"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 寫入數據
for row_idx, row_data in enumerate(data, 2):
    cycle_num = row_data["cycle_num"]
    fill = gray_fill if cycle_num % 2 == 0 else white_fill

    ws.cell(row=row_idx, column=1, value=row_data["週期"]).fill = fill
    ws.cell(row=row_idx, column=2, value=row_data["年份"]).fill = fill
    ws.cell(row=row_idx, column=3, value=row_data["報酬率"]).fill = fill
    ws.cell(row=row_idx, column=4, value=row_data["定義"]).fill = fill
    ws.cell(row=row_idx, column=5, value=row_data["附註"]).fill = fill

    for col in range(1, 6):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col in [2, 3, 4]:  # 年份、報酬率、定義置中
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 設定欄寬
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 45

# 儲存
output_path = os.path.join(SCRIPT_DIR, 'output_經濟週期報酬率分析.xlsx')
wb.save(output_path)
print(f"Excel 已儲存至: {output_path}")
